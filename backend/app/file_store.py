from __future__ import annotations

import csv
import json
import os
import uuid
from pathlib import Path
from openpyxl import load_workbook
from pypdf import PdfReader

BASE_DIR = Path(__file__).resolve().parent.parent / "data" / "uploads"
BASE_DIR.mkdir(parents=True, exist_ok=True)
ALLOWED_EXTENSIONS = {".csv", ".xlsx", ".xlsm", ".pdf", ".txt"}
MAX_FILE_BYTES = 15 * 1024 * 1024
MAX_ROWS_PER_SHEET = 200
MAX_PDF_PAGES = 30
MAX_TEXT_CHARS = 1_000_000


def save_upload(filename: str, content: bytes) -> dict[str, object]:
    original_name = Path(filename or "uploaded_file").name; extension = Path(original_name).suffix.lower()
    if extension not in ALLOWED_EXTENSIONS: raise ValueError(f"Unsupported file type '{extension or 'unknown'}'. Allowed: CSV, XLSX, XLSM, PDF, TXT.")
    if len(content) > MAX_FILE_BYTES: raise ValueError("File is too large. Maximum supported upload size is 15 MB.")
    file_id = uuid.uuid4().hex; stored_name = f"{file_id}{extension}"; destination = BASE_DIR / stored_name; destination.write_bytes(content)
    metadata_path = BASE_DIR / f"{file_id}.meta.json"
    metadata_path.write_text(json.dumps({"file_id":file_id,"filename":original_name,"extension":extension}, indent=2), encoding="utf-8")
    return {"file_id":file_id,"filename":original_name,"extension":extension,"size_bytes":len(content),"analysis_limits":{"max_rows_per_sheet":MAX_ROWS_PER_SHEET,"max_pdf_pages":MAX_PDF_PAGES,"max_text_chars":MAX_TEXT_CHARS}}


def _read_csv(path: Path) -> tuple[str, dict[str, object]]:
    rows=[]; truncated=False
    with path.open("r",encoding="utf-8-sig",newline="") as handle:
        for index,row in enumerate(csv.reader(handle)):
            if index >= MAX_ROWS_PER_SHEET: truncated=True; break
            rows.append(" | ".join(row))
    if truncated: rows.append(f"[TRUNCATED: first {MAX_ROWS_PER_SHEET} rows only]")
    return "\n".join(rows), {"truncated":truncated,"rows_limit":MAX_ROWS_PER_SHEET}


def _read_excel(path: Path) -> tuple[str, dict[str, object]]:
    workbook=load_workbook(path,read_only=True,data_only=True); sections=[]; truncated_sheets=[]
    try:
        for sheet in workbook.worksheets:
            sections.append(f"[SHEET: {sheet.title}]"); truncated=False
            for index,row in enumerate(sheet.iter_rows(values_only=True)):
                if index >= MAX_ROWS_PER_SHEET: truncated=True; break
                sections.append(" | ".join("" if value is None else str(value) for value in row))
            if truncated: sections.append(f"[TRUNCATED: first {MAX_ROWS_PER_SHEET} rows only]"); truncated_sheets.append(sheet.title)
    finally: workbook.close()
    return "\n".join(sections), {"truncated":bool(truncated_sheets),"truncated_sheets":truncated_sheets,"rows_limit":MAX_ROWS_PER_SHEET}


def _read_pdf(path: Path) -> tuple[str, dict[str, object]]:
    reader=PdfReader(str(path)); sections=[]
    for index,page in enumerate(reader.pages[:MAX_PDF_PAGES]): sections.append(f"[PAGE {index+1}]\n{page.extract_text() or ''}")
    truncated=len(reader.pages)>MAX_PDF_PAGES
    if truncated: sections.append(f"[TRUNCATED: first {MAX_PDF_PAGES} pages of {len(reader.pages)} pages]")
    return "\n".join(sections), {"truncated":truncated,"pages_analyzed":min(len(reader.pages),MAX_PDF_PAGES),"total_pages":len(reader.pages)}


def _metadata(file_id: str) -> dict[str, object]:
    path=BASE_DIR / f"{file_id}.meta.json"
    try:
        value=json.loads(path.read_text(encoding="utf-8")); return value if isinstance(value,dict) else {}
    except (FileNotFoundError,json.JSONDecodeError,OSError): return {}


def read_file(file_id: str) -> dict[str, object]:
    if not file_id or Path(file_id).name != file_id or "/" in file_id or "\\" in file_id: raise ValueError("Invalid file id.")
    meta=_metadata(file_id); extension=str(meta.get("extension") or "").lower()
    if extension not in ALLOWED_EXTENSIONS:
        matches=[p for p in BASE_DIR.glob(f"{file_id}.*") if not p.name.endswith(".meta.json")]
        if not matches: raise FileNotFoundError(f"Uploaded file '{file_id}' was not found.")
        path=matches[0]; extension=path.suffix.lower(); original_name=f"{file_id}{extension}"
    else: path=BASE_DIR / f"{file_id}{extension}"; original_name=str(meta.get("filename") or f"{file_id}{extension}")
    if not path.exists(): raise FileNotFoundError(f"Uploaded file '{file_id}' was not found.")
    if extension==".csv": text,analysis_metadata=_read_csv(path)
    elif extension in {".xlsx",".xlsm"}: text,analysis_metadata=_read_excel(path)
    elif extension==".pdf": text,analysis_metadata=_read_pdf(path)
    elif extension==".txt":
        raw=path.read_text(encoding="utf-8-sig"); text=raw[:MAX_TEXT_CHARS]; analysis_metadata={"truncated":len(raw)>MAX_TEXT_CHARS,"characters_analyzed":len(text),"total_characters":len(raw)}
        if analysis_metadata["truncated"]: text += f"\n[TRUNCATED: first {MAX_TEXT_CHARS} characters only]"
    else: raise ValueError(f"Unsupported stored file type '{extension}'.")
    return {"file_id":file_id,"filename":original_name,"extension":extension,"content":text,"characters":len(text),"analysis_metadata":analysis_metadata}
